#  ___________________________________________________________________________
#
#  Pyomo: Python Optimization Modeling Objects
#  Copyright 2017 National Technology and Engineering Solutions of Sandia, LLC
#  Under the terms of Contract DE-NA0003525 with National Technology and
#  Engineering Solutions of Sandia, LLC, the U.S. Government retains certain
#  rights in this software.
#  This software is distributed under the 3-clause BSD License.
#  ___________________________________________________________________________
# --------------------------------------------------------	
#   MODELO SUBASTA 	DE RECONFIGURACIÓN 2019 
#   DESARROLLADO DE USO LIBRE 
#   Oscar Carreño - Rightside SAS - 2019 - www.rightside.app
# --------------------------------------------------------
from pyomo.environ import *
from pyomo.opt import SolverFactory , SolverStatus, TerminationCondition
from pyomo.core import Constraint
from pyomo.opt import ProblemFormat
from pandas import ExcelWriter
from pandas import ExcelFile
import pandas as pd
from openpyxl import load_workbook
from DatosEntrada import *
import linecache
import sys
import math

print ()
print(" EJECUTANDO MODELO DE SUBASTA DE RECONFIGURACIoN Version 1.0") 
print ()

def PrintException():
    exc_type, exc_obj, tb = sys.exc_info()
    f = tb.tb_frame
    lineno = tb.tb_lineno
    filename = f.f_code.co_filename
    linecache.checkcache(filename)
    line = linecache.getline(filename, lineno, f.f_globals)
    print ('SE ENCONTRo UN ERROR ({}, LINE {} "{}"): {}'.format(filename, lineno, line.strip(), exc_obj))

#DEFINICION DE DICCIONARIOS
Pofertado = {}
Qmax = {}
Qmin = {}

try:

	modelo = ConcreteModel()

	modelo.PLANTAS = Set(initialize=plantas.planta, ordered = True)

	for o in ofertas.index:
		Pofertado[o[1]] = ofertas.precio[o]
		Qmax[o[1]] = ofertas.Qmax[o]
		Qmin[o[1]] = ofertas.Qmin[o]


	modelo.oefCompra = Var(modelo.PLANTAS,domain=NonNegativeIntegers)   	
	modelo.oefnoasig = Var(domain=NonNegativeIntegers)   	
	modelo.binasig = Var(modelo.PLANTAS,domain=Boolean)   											

	def fo_rule(rconf):
		expr  = sum (Pofertado[o[1]]*modelo.oefCompra[o[1]]
						for o in ofertas.index)
		expr  += 1.5*PMCC*modelo.oefnoasig
		return expr

	modelo.FuncionObjetivo = Objective(rule=fo_rule, sense=minimize)

	# La maxima asignacion de OEF de Compra   
	modelo.r1 = ConstraintList()
	modelo.r1.add ( 
		sum (modelo.oefCompra[o[1]]
				for o in ofertas.index )
				<= Qsubastada
		)
	# Balance de asignación de OEF de Compra 
	modelo.r2 = ConstraintList()
	modelo.r2.add ( 
		sum ( modelo.oefCompra[o[1]]
				for o in ofertas.index ) + modelo.oefnoasig >= Qsubastada
		)

	# La restricción asociada a la máxima cantidad de OEF de Compra
	modelo.r3 = ConstraintList()
	for o in ofertas.index:
		modelo.r3.add ( 
			Qmax[o[1]] * modelo.binasig[o[1]] - modelo.oefCompra[o[1]] >= 0
		)

	# La restricción asociada a la mínima cantidad de OEF de Compra 
	modelo.r4 = ConstraintList()
	for o in ofertas.index:
		modelo.r4.add ( 
			Qmin[o[1]] * modelo.binasig[o[1]] - modelo.oefCompra[o[1]] <= 0
		)

	opt = SolverFactory(optimizador) 

	opt.options['ratioGap'] = tolerancia
	opt.options['allowableGap'] = toleranciaABS
	opt.options['sec'] = tiempoLimite
		
	modelo.write("subastaRECONF.lp",io_options={"symbolic_solver_labels":True})

	print ()
	print("Optimizador: " , opt.name) 
	results = opt.solve(modelo,tee=1,logfile ="subastaRECONF.log", keepfiles= 0,symbolic_solver_labels=True) 
	print ()


	if (results.solver.status == SolverStatus.ok) and (results.solver.termination_condition == TerminationCondition.optimal):

		print ()
		print ()
		print(" ----- SE ENCONTRo SOLUCIoN oPTIMA ------")
		print ()
		print ("Funcion Objetivo PROBLEMA ENTERO: ", value(modelo.FuncionObjetivo))
		print ()
		print ("ESCRIBIENDO RESULTADOS EN LA BASE DE DATOS: ")
		print ()

		book = load_workbook(xlFile1)
		writer = pd.ExcelWriter(xlFile1, engine='openpyxl') 
		writer.book = book

		sheet = book["resultadoObjetivo"]

		sheet.cell(row=2, column=1).value = value(modelo.FuncionObjetivo)
		sheet.cell(row=2, column=2).value = modelo.oefnoasig.value

		columnasAsignacion = ["AGENTE","PLANTA","QASIGNADA","BINARIA"]

		out_asignacionOEF = pd.DataFrame(columns=columnasAsignacion)

		fila = 0
		for o in ofertas.index:
			if modelo.oefCompra[o[1]].value >= 0.000001:
				fila += 1
				asignacion = []
				asignacion.append(o[0])
				asignacion.append(o[1])
				asignacion.append(modelo.oefCompra[o[1]].value)
				asignacion.append(modelo.binasig[o[1]].value)
				out_asignacionOEF.loc[fila] = asignacion

		writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
		out_asignacionOEF.to_excel(writer,'resultado',index=False)
		writer.save()

	elif (results.solver.termination_condition == TerminationCondition.infeasible):
		print("EL PROBLEMA ES INFACTIBLE 2")
		book = load_workbook(xlFile1)
		sheet = book["resultadoObjetivo"]
		sheet.cell(row=2, column=1).value = ""
		sheet.cell(row=2, column=2).value = ""

	elif(results.solver.termination_condition == TerminationCondition.unbounded):
		print("EL PROBLEMA ES INFACTIBLE 2")
		book = load_workbook(xlFile1)
		sheet = book["resultadoObjetivo"]
		sheet.cell(row=2, column=1).value = ""
		sheet.cell(row=2, column=2).value = ""

	elif(results.solver.termination_condition == TerminationCondition.maxTimeLimit):
		print()
		print("TERMINo POR TIEMPO LiMITE")
		book = load_workbook(xlFile1)
		sheet = book["resultadoObjetivo"]
		sheet.cell(row=2, column=1).value = ""
		sheet.cell(row=2, column=2).value = ""

	else:
		print ("Solver Status: ",  results.solver.status)
		print("La solucion del problema es: ",results.solver.termination_condition)
		print()
		print("TERMINo EJECUCIoN CON ERRORES")
		book = load_workbook(xlFile1)
		sheet = book["resultadoObjetivo"]
		sheet.cell(row=2, column=1).value = ""
		sheet.cell(row=2, column=2).value = ""

except:
	PrintException()
	print()
	print("TERMINo EJECUCIoN CON ERRORES")
	book = load_workbook(xlFile1)
	sheet = book["resultadoObjetivo"]
	sheet.cell(row=2, column=1).value = ""
	sheet.cell(row=2, column=2).value = ""

